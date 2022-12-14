"""
This is the main entry point for the lunch roulette script.
"""

import argparse
from collections import defaultdict
import logging
import random
import re
import subprocess
import sys
import openpyxl
from contextlib import closing
from datetime import datetime

logger = logging.getLogger(__name__)


def main():
    parser = argparse.ArgumentParser(
        description="Read and write to the supplied XLSX file to match people"
        " for lunch roulette"
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="The path to the XLSX file that stores lunch roulette information",
    )
    parser.add_argument(
        "--out",
        help="If supplied, the input XLSX file will not be overwritten for"
        " changes, and this file will be written instead.  Note that the output"
        " will have all formulas replaced with their data.",
    )
    parser.add_argument(
        "--lunch-date",
        type=lambda x: datetime.strptime(x, r"%Y%m%d").date(),
        help="The date of the lunch we're rouletting for, in YYYYMMDD format",
    )
    action_group = parser.add_mutually_exclusive_group(required=True)
    action_group.add_argument(
        "--roulette",
        action="store_true",
        help="Match people for the lunch date and add a column to the XLSX file"
        " named like match_YYYYMMDD with the results, for review.",
    )
    action_group.add_argument(
        "--send-matches",
        action="store_true",
        help="If specified, emails will be sent for the given lunch date.  This"
        " assumes that the XLSX file has a column named like match_YYYYMMDD for"
        " the given lunch date.  The intention is for the XLSX file to be"
        " filled in and reviewed before sending emails.",
    )
    action_group.add_argument(
        "--send-announcement",
        action="store_true",
        help="Send an announcement email to all users that are subscribed to"
        " lunch roulette.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="If specified, we'll print the commands that we would execute"
        " without actually executing them.  So emails will not be sent, for"
        " example.",
    )
    parser.add_argument(
        "--template",
        help="Path to the Outlook template to use with"
        " lunch-roulette-email.ps1.  This argument is required if --send-emails"
        " or --dry-run-send-emails is specified.",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="If specified, the log level will be DEBUG instead of INFO",
    )

    args = parser.parse_args()
    logging.basicConfig(
        stream=sys.stdout, level=logging.DEBUG if args.debug else logging.INFO
    )

    try:
        with closing(
            # Use data_only so that we use the values as of the last time Excel
            # opened the spreadsheet, instead of the raw formulas.  We can't
            # compute the values of formulas in this script; that's too
            # complicated.
            openpyxl.load_workbook(args.xlsx, data_only=True)
        ) as workbook:
            logger.debug(f"Opened XLSX file {args.xlsx}")
            if args.roulette:
                assert (
                    not args.dry_run
                ), "--dry-run is not supported for --roulette"
                assert (
                    args.lunch_date
                ), "--lunch-date argument is required for --roulette"
                out_filename = args.xlsx
                if args.out:
                    out_filename = args.out
                do_roulette(workbook, args.lunch_date, out_filename)
            elif args.send_matches:
                assert (
                    args.lunch_date
                ), "--lunch-date argument is required for sending matches"
                assert (
                    args.template
                ), "--template argument is required when sending emails"
                send_matches(
                    workbook,
                    args.lunch_date,
                    args.template,
                    dry_run=args.dry_run,
                )
            elif args.send_announcement:
                assert (
                    args.template
                ), "--template argument is required when sending emails"
                send_announcement(
                    workbook,
                    args.template,
                    dry_run=args.dry_run,
                )
    except PermissionError:
        logger.error(
            "Permission error!  Make sure that the XLSX file is not already"
            " open in Excel."
        )
        raise


def do_roulette(workbook, lunch_date, out_filename):
    """
    Do the lunch roulette.
    """
    # Assume that the active worksheet is the only interesting one.  This
    # script wasn't written to account for multiple worksheets.
    worksheet = workbook.active

    columns = parse_worksheet_columns(worksheet)
    logger.debug(f"Parsed columns from the workbook: {columns}")

    users = load_users(
        worksheet,
        columns,
        ["email", "frequency", "cluster", "new_to_cluster", "all_matches"],
    )
    logger.debug(f"Parsed {len(users)} users: {users}")

    # We don't really support frequency at the moment.  We only filter out those
    # users that have a frequency of 0.  We also allow 2 as a placeholder for
    # those students that may be willing to meet more frequently.
    users = {k: v for k, v in users.items() if v["frequency"]}
    assert all([v["frequency"] in [None, 0, 1, 2] for v in users.values()])

    matches = match_users(users)
    logger.debug(f"Matches: {matches}")

    update_worksheet_with_matches(
        worksheet, users, columns, matches, lunch_date
    )
    workbook.save(out_filename)
    logger.info(
        f"Saved lunch roulette for {lunch_date.strftime('%Y-%m-%d')} to"
        f" {out_filename}"
    )


def send_matches(workbook, lunch_date, template_path, dry_run=False):
    """
    Send the lunch roulette match emails.
    """
    # Assume that the active worksheet is the only interesting one.  This
    # script wasn't written to account for multiple worksheets.
    worksheet = workbook.active

    columns = parse_worksheet_columns(worksheet)
    logger.debug(f"Parsed columns from the workbook: {columns}")

    match_column_header = make_match_column_header(lunch_date)
    if match_column_header not in columns:
        raise Exception(
            f"XLSX file doesn't contain {match_column_header} column"
        )

    users = load_users(
        worksheet,
        columns,
        [
            "email",
            "friendly_name",
            "full_name",
            "gender",
            "frequency",
            "all_matches",
            match_column_header,
        ],
    )
    logger.debug(f"Parsed {len(users)} users: {users}")

    send_match_emails(users, lunch_date, template_path, dry_run=dry_run)


def send_announcement(workbook, template_path, dry_run=False):
    """
    Send an announcement to everybody, not only those with matches on a
    particular day.
    """
    # Assume that the active worksheet is the only interesting one.  This
    # script wasn't written to account for multiple worksheets.
    worksheet = workbook.active

    columns = parse_worksheet_columns(worksheet)
    logger.debug(f"Parsed columns from the workbook: {columns}")

    users = load_users(
        worksheet,
        columns,
        [
            "email",
            "friendly_name",
            "frequency",
        ],
    )
    logger.debug(f"Parsed {len(users)} users: {users}")

    send_announcement_emails(users, template_path, dry_run=dry_run)


def parse_worksheet_columns(worksheet):
    """
    Parse the provided workbook to identify the columns that we care about.
    """
    columns = {
        "email": None,
        "frequency": None,
        "friendly_name": None,
        "full_name": None,
        "gender": None,
        "cluster": None,
        "year": None,
        "new_to_cluster": None,
        "all_matches": [],  # This is a special storing all of the match columns
        "first_empty": None,  # This is a placeholder
    }
    required_columns = [
        "email",
        "friendly_name",
        "full_name",
        "gender",
        "cluster",
        "year",
    ]

    # Iterator through the first row, assuming that it contains all of the
    # column headers.
    column_number = 1  # The current column's number
    value = worksheet.cell(row=1, column=column_number).value
    while value:
        logger.debug(f"Column #{column_number} header: {value}")

        # Save all of the match columns.
        if is_match_column_header(value):
            # Match columns can be duplicated, to signify that some people got more than one match.
            if value not in columns:
                columns[value] = []
            columns[value].append(column_number)
            columns["all_matches"].append(column_number)
        elif value in columns and value not in ["first_empty"]:
            columns[value] = column_number

        column_number += 1
        value = worksheet.cell(row=1, column=column_number).value

    assert columns["first_empty"] is None
    columns["first_empty"] = column_number

    for required_column in required_columns:
        if columns[required_column] is None:
            raise Exception(
                f"Worksheet missing required column {required_column}"
            )

    # Remove all None columns from the result.  This simplifies some of the
    # usage elsewhere.
    column_list = list(columns.keys())
    for column in column_list:
        if columns[column] is None:
            del columns[column]

    return columns


def load_users(worksheet, columns, load_columns):
    """
    Load information about each of our users into memory.  Identify each user
    with the corresponding row number in the spreadsheet.  We use email as the
    main key because it's a required field and it's usually useful later on.
    """
    users = {}  # Key is user row number, value is a dictionary with user
    # information
    emails = {}  # Key is email, value is user ID

    columns = {
        k: v for k, v in columns.items() if k in load_columns + ["email"]
    }

    row_number = 2  # Skip the header row (see parse_worksheet_columns)
    value = worksheet.cell(row=row_number, column=columns["email"]).value
    match_columns = set()
    while value:
        users[row_number] = {
            "email": value,
            "id": row_number,
        }
        emails[value] = row_number
        user = users[row_number]
        for column in columns:
            if columns[column] is None:
                # Transmit the None so future calls can dereference the
                # column easily, without needing to first check if the key
                # exists.
                user[column] = None
            elif column == "all_matches" or is_match_column_header(column):
                # Match columns are special and contain a list of all of columns.
                if column not in user:
                    user[column] = []
                    match_columns.add(column)
                for match_col in columns[column]:
                    value = worksheet.cell(
                        row=row_number, column=match_col
                    ).value
                    if value is not None:
                        user[column].append(value)
            else:
                user[column] = worksheet.cell(
                    row=row_number, column=columns[column]
                ).value

        row_number += 1
        value = worksheet.cell(row=row_number, column=columns["email"]).value

    # Convert the emails from the match columns into IDs.
    for id, user in users.items():
        for match_column in match_columns:
            if match_column in user:
                match_ids = []
                for match_email in user[match_column]:
                    match_ids.append(emails[match_email])
                user[match_column] = match_ids

    return users


def make_match_column_header(lunch_date):
    return f"match_{lunch_date.strftime(r'%Y%m%d')}"


def is_match_column_header(header):
    return re.match(r"^match_\d{8}$", header) is not None


def match_users(users):
    """
    Return a list of tuples representing each of the matches.
    """
    scores = {}
    users_by_score = defaultdict(set)

    def score_match(user_a, user_b):
        # Avoid matching people that have already been matched.
        if user_a["id"] in user_b["all_matches"]:
            user_b["id"] in user_a["all_matches"]
            return -1

        if user_a["new_to_cluster"] or user_b["new_to_cluster"]:
            return 2 if user_a["cluster"] == user_b["cluster"] else 1

        # For everyone else, prefer matching with someone outside of their
        # cluster.
        return 1 if user_a["cluster"] != user_b["cluster"] else 0

    def remove_all_matches_for_user(user_id):
        for other_user_id in users:
            if user_id == other_user_id:
                continue
            pair = (
                min(user_id, other_user_id),
                max(user_id, other_user_id),
            )
            if pair in scores:
                users_by_score[scores[pair]].remove(pair)
                del scores[pair]

    # Make a 2D map of two user IDs to their "score", which indicates how likely
    # we should be to pair these two up.
    for first_user_id in users:
        for second_user_id in users:
            if first_user_id >= second_user_id:
                continue

            score = score_match(users[first_user_id], users[second_user_id])
            pair = (first_user_id, second_user_id)
            scores[pair] = score
            users_by_score[score].add(pair)

    matches = []
    sorted_scores = sorted(users_by_score.keys(), reverse=True)
    for score in sorted_scores:
        # Randomly match one pair at a time until we run out of pairs.
        while users_by_score[score]:
            match = random.choice(list(users_by_score[score]))
            matches.append(match)

            # Clear this match from everywhere
            remove_all_matches_for_user(match[0])
            remove_all_matches_for_user(match[1])

    return matches


def update_worksheet_with_matches(
    worksheet, users, columns, matches, lunch_date
):
    """
    Update the XLSX workbook with each person's match.  The matches are a list
    of tuples, with each user identified by their row number in the spreadsheet.
    """

    # First make the new match header, at the end of the sheet so as not to
    # invalidate any of our other indices.
    match_column = columns["first_empty"]
    worksheet.cell(row=1, column=match_column).value = make_match_column_header(
        lunch_date
    )

    # Now go through each match and write it into the spreadsheet.
    for match in matches:
        emails = (users[match[0]]["email"], users[match[1]]["email"])
        worksheet.cell(row=match[0], column=match_column).value = emails[1]
        worksheet.cell(row=match[1], column=match_column).value = emails[0]


def send_match_emails(users, lunch_date, template_path, dry_run=False):
    """
    Send emails to each person about their match.
    """
    # Make a dictionary mapping the users' email address back to their row
    # numbers.  We'll use this for getting the match information.
    users_by_id = {v["id"]: v for v in users.values()}
    match_column_header = make_match_column_header(lunch_date)

    # Send emails serially, because I doubt that Powershell and Outlook support
    # sending emails in parallel.
    pretty_date = lunch_date.strftime("%A %B %d, %Y")
    send_failures = []  # Tracks the send failures that we encountered.
    for user in users.values():
        if user[match_column_header] and user["frequency"] > 0:
            matches = [
                users_by_id[match_id] for match_id in user[match_column_header]
            ]

            def join_emails(matches):
                # https://www.rfc-editor.org/rfc/rfc6068#section-2 says that commas
                # are valid delimiters.  This may not be implemented in all email
                # clients, though.
                return ",".join(m["email"] for m in matches)

            def join_names(matches, key):
                if len(matches) <= 2:
                    return " and ".join(m[key] for m in matches)
                else:
                    return (
                        ", ".join(m[key] for m in matches[:-1])
                        + f", and {matches[-1][key]}"
                    )

            def join_genders(matches):
                return matches[0]["gender"] if len(matches) == 1 else "plural"

            args = [
                "powershell.exe",
                ".\\lunch-roulette-email.ps1",
                "-email",
                f"'{user['email']}'",
                "-template",
                f"'{template_path}'",
                "-replacements",
                "@{"
                f"'VarFriendlyName'='{user['friendly_name']}'"
                f"; 'VarLunchDate'='{pretty_date}'"
                f"; 'VarOtherEmail'='{join_emails(matches)}'"
                f"; 'VarOtherFriendlyName'='{join_names(matches, 'friendly_name')}'"
                f"; 'VarOtherFullName'='{join_names(matches, 'full_name')}'"
                f"; 'VarOtherGender'='{join_genders(matches)}'"
                "}",
            ]
            logger.info(f"Sending email to {user['email']}...")

            if dry_run:
                print(" ".join(args))
            else:
                completed_process = subprocess.run(args)
                if completed_process.returncode != 0:
                    logger.error(f"Failed to send email to {user['email']}")
                    send_failures.append(user)

    if send_failures:
        logger.error(
            "Failed to send emails to the following users:"
            + "\n  ".join([u["email"] for u in send_failures])
        )


def send_announcement_emails(users, template_path, dry_run=False):
    # Send emails serially, because I doubt that Powershell and Outlook support
    # sending emails in parallel.
    send_failures = []  # Tracks the send failures that we encountered.
    for user in users.values():
        if user["frequency"] > 0:
            args = [
                "powershell.exe",
                ".\\lunch-roulette-email.ps1",
                "-email",
                f"'{user['email']}'",
                "-template",
                f"'{template_path}'",
                "-replacements",
                "@{" f"'VarFriendlyName'='{user['friendly_name']}'" "}",
            ]
            logger.info(f"Sending email to {user['email']}...")

            if dry_run:
                print(" ".join(args))
            else:
                completed_process = subprocess.run(args)
                if completed_process.returncode != 0:
                    logger.error(f"Failed to send email to {user['email']}")
                    send_failures.append(user)

    if send_failures:
        logger.error(
            "Failed to send emails to the following users:"
            + "\n  ".join([u["email"] for u in send_failures])
        )


if __name__ == "__main__":
    main()
